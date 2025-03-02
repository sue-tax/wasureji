'''
Created on 2024/12/14
@author: sue-t
'''

from main.sequence import Sequence

import openpyxl
# import glob
import datetime
import os
import sys
import json
import win32com.client
import TkEasyGUI as eg

'''
検索
select base.document, input.by, output.delivery from base
CROSS JOIN input ON base.file=input.file
CROSS JOIN output ON base.file=output.file
'''

# TODO Excel出力

class wasureji_utility(object):
    font_ = ("ＭＳゴシック")
    
    layout_utility = [
        [eg.Button("サーバー終了", key="-kill-")],
        [eg.Label("")],

        [eg.Label("ＳＱＬ実行", font=(32)),
            eg.Button("実行", key="-sql_exe-"),
            eg.Label("　　　"),
            eg.Button("コミット", key="-sql_commit-"),
            eg.Label("　"),
            eg.Button("ロールバック", key="-sql_rollback-")],
        [eg.Label("ＳＱＬ文")],
        [eg.Multiline(
                "",
                key="-sql-",
                size=(120, 5))],

        [eg.Label("結果"),
            eg.Label("　　　　　　　　　　　　"),
            eg.Button("Excel", key="-excel-")
            ],
        [eg.Multiline(
                "",
                key="-result-",
                size=(120, 10))],
        [eg.Label("")],
        
        # 検索等
        [eg.Label("検索等", font=(32)),
            eg.Label("　　"),
            eg.Button("表示", key="-sql_find-"),
            eg.Label("　　　"),
            # eg.Button("置換", key="-sql_update-"),
            eg.Label("　　　　　　　　　　　"),
            # eg.Button("削除", key="-sql_delete-"),
            eg.Label("　　　　　　　　　　　　　　"),
            eg.Button("再読込", key="-reload-")    # Combo再設定
            ],
        [eg.Label("　　　　　　　　"),
            eg.Label("表示"),
            eg.Label("　　　　"),
            # eg.Label("置換　　　　　　　"),
            eg.Label("　　　　　"),
            # eg.Label("削除　　"),
            eg.Label("検索条件　　　　　　　　　　")
            ],
            
        [
            eg.Label("ファイル名", font=font_),
            eg.Label("　　　"),
            eg.Checkbox(key="-cb_select_file-"),
            eg.Label("　　"),
            # eg.Checkbox(key="-cb_update_file-"),
            # eg.Input("", width=20, key="-inp_update_file-"),
            # eg.Label("　　　　　　　"),
            eg.Label("　　　　　　"),
            eg.Checkbox(key="-cb_where_file-"),
            eg.Input("", width=20, key="-inp_where_file-")
            ],
        [eg.Label("基本＿"),
            eg.Label("　　　　　　　　　"),
            # eg.Label("置換"),
            # eg.Checkbox(key="-cb_update_file-"),
            # eg.Label("　　　　　　　　　　　"),
            # eg.Label("削除"),
            # eg.Checkbox(key="-cb_delete_file-")
            ],
        [
            eg.Label("書類名＿", font=font_),# font=font_),
            eg.Label("　　　"),
            eg.Checkbox(key="-cb_select_document-"),
            eg.Label("　　　"),
            # eg.Checkbox(key="-cb_update_document-"),
            # eg.Combo("", width=20, key="-inp_update_document-"),
            eg.Label("　　　　　"),
            eg.Checkbox(key="-cb_where_document-"),
            eg.Combo("", key="-inp_where_document-")
            ],
        [
            eg.Label("顧客名＿", font=font_),# font=font_),
            eg.Label("　　　"),
            eg.Checkbox(key="-cb_select_customer-"),
            eg.Label("　　　"),
            # eg.Checkbox(key="-cb_update_customer-"),
            # eg.Combo("", width=20, key="-inp_update_customer-"),
            eg.Label("　　　　　"),
            eg.Checkbox(key="-cb_where_customer-"),
            eg.Combo("", key="-inp_where_customer-")
            ],
        [
            eg.Label("区分名＿", font=font_),# font=font_),
            eg.Label("　　　"),
            eg.Checkbox(key="-cb_select_section-"),
            eg.Label("　　　"),
            # eg.Checkbox(key="-cb_update_section-"),
            # eg.Combo("", width=20, key="-inp_update_section-"),
            eg.Label("　　　　　"),
            eg.Checkbox(key="-cb_where_section-"),
            eg.Combo("", key="-inp_where_section-"),
            eg.Label("　　　　　　　　　　　　　　　　　")
            ],

        [eg.Label("受取り"),
            eg.Label("　　　　　　　　　"),
            # eg.Label("置換"),
            # eg.Checkbox(key="-cb_update_input-"),
            # eg.Label("　　　　　　　　　　　"),
            # eg.Label("削除"),
            # eg.Checkbox(key="-cb_delete_input-")
            ],
        [
            eg.Label("いつ＿＿", font=font_),
            eg.Label("　　　"),
            eg.Checkbox(key="-cb_select_indate-"),
            eg.Label("　　　"),
            # eg.Checkbox(key="-cb_update_indate-"),
            # eg.Input("", width=20, key="-inp_update_indate-"),
            # eg.Label("　　　　　　　"),
            eg.Label("　　　　　 "),
            eg.Checkbox(key="-cb_where_indate-"),
            eg.Input("", key="-inp_where_indate-"),
            eg.Label("　～　"),
            eg.Checkbox(key="-cb_where_indateto-"),
            eg.Input("", key="-inp_where_indateto-")
            ],
        [
            eg.Label("誰から＿", font=font_),
            eg.Label("　　　"),
            eg.Checkbox(key="-cb_select_inorigin-"),
            eg.Label("　　　"),
            # eg.Checkbox(key="-cb_update_inorigin-"),
            # eg.Combo("", width=20, key="-inp_update_inorigin-"),
            eg.Label("　　　　　"),
            eg.Checkbox(key="-cb_where_inorigin-"),
            eg.Combo("", key="-inp_where_inorigin-")
            ],
        [
            eg.Label("何で＿＿", font=font_),
            eg.Label("　　　"),
            eg.Checkbox(key="-cb_select_inby-"),
            eg.Label("　　　"),
            # eg.Checkbox(key="-cb_update_inby-"),
            # eg.Combo("", width=20, key="-inp_update_inby-"),
            eg.Label("　　　　　"),
            eg.Checkbox(key="-cb_where_inby-"),
            eg.Combo("", key="-inp_where_inby-")
            ],

        [eg.Label("引渡し"),
            eg.Label("　　　　　　　　　"),
            # eg.Label("置換"),
            # eg.Checkbox(key="-cb_update_output-"),
            # eg.Label("　　　　　　　　　　　"),
            # eg.Label("削除"),
            # eg.Checkbox(key="-cb_delete_output-")
            ],
        [
            eg.Label("いつ＿＿", font=font_),
            eg.Label("　　　"),
            eg.Checkbox(key="-cb_select_outdate-"),
            eg.Label("　　　"),
            # eg.Checkbox(key="-cb_update_outdate-"),
            # eg.Input("", width=20, key="-inp_update_outdate-"),
            # eg.Label("　　　　　　　"),
            eg.Label("　　　　　 "),
            eg.Checkbox(key="-cb_where_outdate-"),
            eg.Input("", key="-inp_where_outdate-"),
            eg.Label("　～　"),
            eg.Checkbox(key="-cb_where_outdateto-"),
            eg.Input("", key="-inp_where_outdateto-")
            ],
        [
            eg.Label("誰へ＿＿", font=font_),
            eg.Label("　　　"),
            eg.Checkbox(key="-cb_select_outdelivery-"),
            eg.Label("　　　"),
            # eg.Checkbox(key="-cb_update_outdelivery-"),
            # eg.Combo("", width=20, key="-inp_update_outdelivery-"),
            eg.Label("　　　　　"),
            eg.Checkbox(key="-cb_where_outdelivery-"),
            eg.Combo("", key="-inp_where_outdelivery-")
            ],
        [
            eg.Label("何で＿＿", font=font_),
            eg.Label("　　　"),
            eg.Checkbox(key="-cb_select_outby-"),
            eg.Label("　　　"),
            # eg.Checkbox(key="-cb_update_outby-"),
            # eg.Combo("", width=20, key="-inp_update_outby-"),
            eg.Label("　　　　　"),
            eg.Checkbox(key="-cb_where_outby-"),
            eg.Combo("", key="-inp_where_outby-")
            ],

        [eg.Label("")],
        [eg.Button("終了", key="-exit-")]
    ]

    def combo_set(self):
        list_doc = self.seq.send_ask_document()
        list_cust = self.seq.send_ask_customer()
        list_sect = self.seq.send_ask_section()
        list_in_origin = self.seq.send_ask_in_origin()
        list_in_by = self.seq.send_ask_in_by()
        list_out_delivery = self.seq.send_ask_out_delivery()
        list_out_by = self.seq.send_ask_out_by()
        self.window["-inp_where_document-"].set_values(list_doc)
        self.window["-inp_where_customer-"].set_values(list_cust)
        self.window["-inp_where_section-"].set_values(list_sect)
        self.window["-inp_where_inorigin-"].set_values(list_in_origin)
        self.window["-inp_where_inby-"].set_values(list_in_by)
        self.window["-inp_where_outdelivery-"].set_values(list_out_delivery)
        self.window["-inp_where_outby-"].set_values(list_out_by)
        # self.window["-inp_update_document-"].set_values(list_doc)
        # self.window["-inp_update_customer-"].set_values(list_cust)
        # self.window["-inp_update_section-"].set_values(list_sect)
        # self.window["-inp_update_inorigin-"].set_values(list_in_origin)
        # self.window["-inp_update_inby-"].set_values(list_in_by)
        # self.window["-inp_update_outdelivery-"].set_values(list_out_delivery)
        # self.window["-inp_update_outby-"].set_values(list_out_by)

    def __init__(self, host, port, directory,
            excel_file): #, excel_sheet):
        self.host = host
        self.port = port
        self.directory = directory
        self.excel_file = excel_file
        # self.excel_sheet = excel_sheet

    def start(self):
        self.seq = Sequence(self.host, self.port)
        str_msg = self.seq.ping()
        if str_msg != None:
            eg.popup_error(str_msg,
                    "サーバーが起動していない")
            return -1
        self.window = eg.Window("wasurejiユーティリティ",
                self.layout_utility)
        self.window["-result-"].set_readonly(True)
        # self.window["-excel-"].set_disabled(True)
        # self.window["-sql_update-"].set_disabled(True)
        # self.window["-sql_delete-"].set_disabled(True)
        self.combo_set()
        while True:
            event, _values = self.window.read()
            if event == "-kill-":
                answer = eg.confirm(
                        "wasurejiサーバーの終了後、終了するので良いですか？"
                        "")
                if answer:
                    self.seq.send_kill()
                    break
                continue
            if event == "-sql_exe-":
                str_sql = self.window["-sql-"].get()
                if str_sql == "":
                    continue
                # print(str_sql)
                msg_sql = str_sql.rstrip('\n')
                # print(msg_sql)
                str_result = self.seq.send_execute_sql(msg_sql)
                # print(str_result)
                self.window["-result-"].set_text(str_result)
                continue

            if event == "-excel-":
                self.to_excel()
                continue
                
            if event == "-sql_find-":
                self.find()
                continue
            if event == "-sql_update-":
                self.update()
                continue
            if event == "-reload-":
                self.combo_set()
                continue

            if event == "-exit-" or event == "WINDOW_CLOSED":
                answer = eg.confirm(
                        "終了して良いですか？",
                        "")
                if answer:
                    break
                continue
            if event == "WINDOW_CLOSED":
                break

        self.window.close()
    
    
    def to_excel(self):
        str_result = self.window["-result-"].get()
        if (str_result[0] != "[") or (str_result[-1] != "]"):
            eg.popup_auto_close(
                "検索結果がありません。",
                "Information")
            return
        str_ = str_result[1:-1]
        
        # excel_trunc = "wasureji"
        # excel_file_name = "wasureji.xlsx"
        # num_file = 2
        # files = glob.glob("*.xlsx")
        # while excel_file_name in files:
        #     # excel_trunc += "_"
        #     excel_file_name = excel_trunc + "_" + str(num_file) + ".xlsx"
        #     num_file += 1
        
        # base_file = os.path.splitext(self.excel_file)
        now = datetime.datetime.now()
        file_name = self.excel_file + '_' \
                + now.strftime('%Y%m%d%H%M%S') + '.xlsx'
        excel_file_name = os.path.join(
                self.directory, file_name)
        # print(excel_file_name)
        create_file = openpyxl.Workbook()
        create_file.save(excel_file_name)

        excel_file = openpyxl.load_workbook(excel_file_name)
        # sheet_list = excel_file.sheetnames
        # sheet_name = 'sheet_wasureji'
        # sheet_name = self.excel_sheet
        # if sheet_name in sheet_list:
        #     pass
        # else:
        #     excel_file.create_sheet(title=sheet_name)
        # sheet = excel_file[sheet_name]
        # excel_file.active = sheet
        sheet = excel_file.active
        
        str_sql = self.window["-sql-"].get()
        sheet.cell(row=1, column=1, value=str_sql)
        start_row = 2
        start_col = 1
        list_line = str_.split('\n')
        for y, row in enumerate(list_line):
            list_ = row[1:-2].split(',')
            for x, cell_ in enumerate(list_):
                sheet.cell(row=start_row + y,
                           column=start_col + x,
                           value=cell_[1:])
        excel_file.save(excel_file_name)
        
    def find(self):
        list_sql = ["SELECT "]
        if self.window["-cb_select_file-"].get():
            list_sql.append("base.file,")
        if self.window["-cb_select_document-"].get():
            list_sql.append("base.document,")
        if self.window["-cb_select_customer-"].get():
            list_sql.append("base.customer,")
        if self.window["-cb_select_section-"].get():
            list_sql.append("base.section,")

        if self.window["-cb_select_indate-"].get():
            list_sql.append("input.date,")
        if self.window["-cb_select_inorigin-"].get():
            list_sql.append("input.origin,")
        if self.window["-cb_select_inby-"].get():
            list_sql.append("input.by,")

        if self.window["-cb_select_outdate-"].get():
            list_sql.append("output.date,")
        if self.window["-cb_select_outdelivery-"].get():
            list_sql.append("output.delivery,")
        if self.window["-cb_select_outby-"].get():
            list_sql.append("output.by,")

        if len(list_sql) == 1:
            eg.popup_auto_close(
                "表示が、どれも選択されていません。",
                "Error")
            return
        atom_ = list_sql.pop(-1)
        list_sql.append(atom_[:-1])
        list_sql.append(" FROM base ")
        list_sql.append("LEFT OUTER JOIN input ON base.file=input.file ")
        list_sql.append("LEFT JOIN output ON base.file=output.file ")
        where_ = self.create_where()
        list_sql.append(where_)
        # print(list_sql)
        str_sql = ''.join(list_sql)
        # print(str_sql)
        str_result = self.seq.send_execute_sql(str_sql)
        # print(str_result)
        self.window["-sql-"].set_text(str_sql)
        self.window["-result-"].set_text(str_result)

    def update(self):
        
        # update base SET customer="J" FROM base LEFT OUTER JOIN input ON base.file=input.file LEFT JOIN output ON base.file=output.file WHERE base.customer="ジェイオー"
        # ambiguous column name: base.customer
        
        # update base SET customer="J" FROM base as bs LEFT OUTER JOIN input ON bs.file=input.file LEFT JOIN output ON bs.file=output.file WHERE bs.customer="ジェイオー"        
        # 成功のようだ
        
        # TODO
        
        str_sql = "UPDATE base,input,output SET "
        list_sql = []
        if self.window["-cb_update_file-"].get():
            list_sql.append('base.file="' +
                    self.window["-inp_update_file-"].get() +
                    '",')
        # if self.window["-cb_select_document-"].get():
        #     list_sql.append("base.document,")
        # if self.window["-cb_select_customer-"].get():
        #     list_sql.append("base.customer,")
        # if self.window["-cb_select_section-"].get():
        #     list_sql.append("base.section,")
        #
        # if self.window["-cb_select_indate-"].get():
        #     list_sql.append("in.date,")
        # if self.window["-cb_select_inorigin-"].get():
        #     list_sql.append("in.origin,")
        # if self.window["-cb_select_inby-"].get():
        #     list_sql.append("in.by,")
        #
        # if self.window["-cb_select_outdate-"].get():
        #     list_sql.append("out.date,")
        # if self.window["-cb_select_outdelivery-"].get():
        #     list_sql.append("out.delivery,")
        # if self.window["-cb_select_outby-"].get():
        #     list_sql.append("out.by,")


        list_sql.append("LEFT OUTER JOIN input ON base.file=input.file ")
        list_sql.append("LEFT JOIN output ON base.file=output.file ")
        where_ = self.create_where()
        list_sql.append(where_)
        print(list_sql)
        str_sql = ''.join(list_sql)
        print(str_sql)
        str_result = self.seq.send_execute_sql(str_sql)
        # print(str_result)
        self.window["-result-"].set_text(str_result)
    
    def create_where(self):
        # list_sql = [ "WHERE " ]
        list_sql = []
        if self.window["-cb_where_file-"].get():
            list_sql.append('base.file="' +
                    self.window["-inp_where_file-"].get()+'"')
        if self.window["-cb_where_document-"].get():
            list_sql.append('base.document="' +
                    self.window["-inp_where_document-"].get()+'"')
        if self.window["-cb_where_customer-"].get():
            list_sql.append('base.customer="' +
                    self.window["-inp_where_customer-"].get()+'"')
        if self.window["-cb_where_section-"].get():
            list_sql.append('base.section="' +
                    self.window["-inp_where_section-"].get()+'"')

        if self.window["-cb_where_indate-"].get():
            if not self.window["-cb_where_indateto-"].get():
                list_sql.append('input.date="' +
                        self.window["-inp_where_indate-"].get()+'"')
            else:
                list_sql.append('input.date>=' +
                        self.window["-inp_where_indate-"].get())
                list_sql.append('input.date<=' +
                        self.window["-inp_where_indateto-"].get())
        if self.window["-cb_where_inorigin-"].get():
            list_sql.append('input.origin="' +
                    self.window["-inp_where_inorigin-"].get()+'"')
        if self.window["-cb_where_inby-"].get():
            list_sql.append('input.by="' +
                    self.window["-inp_where_inby-"].get()+'"')
                
        if self.window["-cb_where_outdate-"].get():
            if not self.window["-cb_where_outdateto-"].get():
                list_sql.append('output.date="' +
                        self.window["-inp_where_outdate-"].get()+'"')
            else:
                list_sql.append('output.date>=' +
                        self.window["-inp_where_outdate-"].get())
                list_sql.append('output.date<=' +
                        self.window["-inp_where_outdateto-"].get())
        if self.window["-cb_where_outdelivery-"].get():
            list_sql.append('output.delivery="' +
                    self.window["-inp_where_outdelivery-"].get()+'"')
        if self.window["-cb_where_outby-"].get():
            list_sql.append('output.by="' +
                    self.window["-inp_where_outby-"].get()+'"')
        
        if len(list_sql) == 0:
            return ""
        str_where = "WHERE " + " AND ".join(list_sql)
        # print(str_where)
        return str_where

if __name__ == '__main__':
    try:
        objShell = win32com.client.Dispatch("WScript.Shell")
        dir_name = objShell.SpecialFolders("SENDTO")
        json_file_name = os.path.join(dir_name, 'wasureji.json')
        json_file = open(json_file_name, 'r')
        json_dict = json.load(json_file)
    except Exception as e:
        eg.popup_error(
                "設定ファイル{}が読めません".format(
                        json_file_name),
                "wasureji_utility")
        sys.exit(-1)
    ut = wasureji_utility(
            json_dict["host"], json_dict["port"],
            json_dict["directory"],
            json_dict["excel_file"])
            # json_dict["excel_sheet"])
    ut.start()
    sys.exit(0)
